"""
build_master.py — Consolidates every data source into one professional Excel workbook.

Sheets:
  1. Summary         — Stats at a glance
  2. Master          — Every unique lead, all sources, deduplicated
  3. Showroom Lists  — UK KBB showrooms (Companies House / manual sources)
  4. Job Leads       — Companies hiring CAD/designer roles (job board data)
  5. Golden Leads    — High-value / existing customers (hidden tab, special access)
  6. Missing Data    — Leads with gaps: no email, phone or website
  7. Team Input      — Clean template: team adds company name, agent fills the rest
"""
import os, re, glob, warnings
from collections import Counter
from io import StringIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore")

BASE     = os.path.dirname(os.path.abspath(__file__))
OLD_DIR  = os.path.join(BASE, "Old")
JOB_CSV  = os.path.join(BASE, "..", "vw-job-agent", "state", "new_jobs.csv")

# ── Design tokens ──────────────────────────────────────────────────────────────
GOLD      = "C9A84C"
DARK      = "1A1A2E"
MID       = "16213E"
LIGHT_BG  = "F7F4EE"
ALT_BG    = "FFFFFF"
HOT_CLR   = "E74C3C"
WARM_CLR  = "E67E22"
WATCH_CLR = "3498DB"
GREEN     = "27AE60"
PURPLE    = "8E44AD"
GREY      = "95A5A6"
RED_LIGHT = "FADBD8"
AMBER     = "FDEBD0"

# ── Unified lead schema ────────────────────────────────────────────────────────
LEAD_COLS = [
    "Company Name", "Owner / Director", "Category", "Brand(s)",
    "Country", "City", "Postcode", "Full Address",
    "Phone", "Email", "Website",
    "LinkedIn (Owner)", "LinkedIn (Company)", "Instagram", "Facebook",
    "Outreach Status", "Email Sent Date", "Follow-up Date",
    "Response", "Notes", "Sales Person", "Source",
    "Tier", "Score",
]

COL_WIDTHS = {
    "Company Name": 34, "Owner / Director": 24, "Category": 13,
    "Brand(s)": 20, "Country": 11, "City": 15, "Postcode": 10,
    "Full Address": 38, "Phone": 18, "Email": 30, "Website": 30,
    "LinkedIn (Owner)": 30, "LinkedIn (Company)": 30,
    "Instagram": 26, "Facebook": 26,
    "Outreach Status": 18, "Email Sent Date": 15, "Follow-up Date": 15,
    "Response": 16, "Notes": 35, "Sales Person": 16, "Source": 22,
    "Tier": 8, "Score": 7,
}

# Job leads have a different schema
JOB_COLS = [
    "Found Date", "Tier", "Score", "Job Title", "Company", "Showroom Name",
    "Location", "Salary", "Is Recruiter", "Opening Line", "Job URL", "Source",
]
JOB_WIDTHS = {
    "Found Date": 12, "Tier": 8, "Score": 7, "Job Title": 30, "Company": 28,
    "Showroom Name": 28, "Location": 18, "Salary": 16, "Is Recruiter": 12,
    "Opening Line": 40, "Job URL": 38, "Source": 16,
}

# ── Style helpers ──────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def border(color="D0C8B0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", wrap=False, v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header_cell(ws, row, col, value, bg=DARK, fg=GOLD, sz=10, bold=True, center=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg); c.font = font(bold=bold, color=fg, size=sz)
    c.alignment = align("center" if center else "left")
    c.border = border()
    return c

def data_cell(ws, row, col, value, bg=LIGHT_BG, fg="111111", wrap=False, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg); c.font = font(bold=bold, color=fg)
    c.alignment = align(wrap=wrap)
    c.border = border()
    return c

def set_col_widths(ws, cols, widths_map):
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths_map.get(col, 18)

def add_status_dropdown(ws, col_idx, start_row, end_row,
                        values='"Not Contacted,Email Sent,Followed Up,Replied,Connected,Not Interested,Unsubscribed"'):
    dv = DataValidation(type="list", formula1=values,
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    col_letter = get_column_letter(col_idx)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"

# ── Data cleaning ──────────────────────────────────────────────────────────────
JUNK = {"", "nan", "NaN", "None", "none", "N/A", "n/a", "-", "NA",
        "Unavailable", "Not known", "NaT", "nat"}

def _c(row, *keys):
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip() not in JUNK:
            return str(v).strip()
    return None

def _norm(name):
    return re.sub(r"[^a-z0-9]", "", (name or "").lower())

# ── Row → unified lead ─────────────────────────────────────────────────────────
def row_to_lead(row, source, country=None, category=None):
    r = {c: None for c in LEAD_COLS}
    row = {str(k): v for k, v in row.items()}

    r["Company Name"]       = _c(row, "Company name","KITCHEN RETAILERS","BATHROOM RETAILERS",
                                  "COMPANY","Name","Kitchens Retailers Name",
                                  "Showroom/Business Name"," Company Name                      ",
                                  "company_name","lead_name","COMPANY NAME")
    r["Owner / Director"]   = _c(row, "Owner Name","COMPANY OWNER","COMPANY OWNER ","Owner",
                                  "owner_name","OWNER/DIRECTOR  NAME","Company owner ","owner",
                                  "Owner/Director Name")
    r["Brand(s)"]           = _c(row, "BRAND","BRAND.1","BRAND.2","Brand",
                                  "Kitchen Brands Sold","Brands","brand"," Brand")
    r["Country"]            = _c(row, "COUNTRY","country") or country
    r["City"]               = _c(row, "City","CITY","city")
    r["Postcode"]           = _c(row, "Postcode","POST CODE","POSTCODE","Post code","postcode")
    r["Full Address"]       = _c(row, "Address","ADDRESS","Full Address",
                                  "Street Address","address")
    r["Phone"]              = _c(row, "Phone Numbers","CONTACT NO.","CONTACT","Phone",
                                  "TELEPHONE ","contact_number","Phone number",
                                  "Apollo phone","APOLLO PHONE","Mobile","phone","CONTACT NO")
    r["Email"]              = _c(row, "Email ","EMAIL ID","EMAIL","Email","E-MAIL",
                                  "Apollo Email","APOLLO EMAIL","email","Email ID")
    r["Website"]            = _c(row, "Websaite","WEBSITE","Website ","WEBSITE ","website","Website")
    r["LinkedIn (Owner)"]   = _c(row, "Linkdin","LINKED IN ","LINKED IN","Owner Linkedin",
                                  "Owner's Linkedin","OWNER'S LINKEDIN","Linkedin id",
                                  "linkedin_owner","owner_linkedin","LINKEDIN")
    r["LinkedIn (Company)"] = _c(row, "Company Linkedin page","Linkedin page","linkedin_company")
    r["Instagram"]          = _c(row, "instagram ","INSTAGRAM","Instagram","instagram")
    r["Facebook"]           = _c(row, "facebook","FACEBOOK","Facebook")
    r["Outreach Status"]    = _c(row, "outreach_status","Did we contacted them",
                                  "COMMUNICATION","Showroom visit","showroom  visit")
    r["Email Sent Date"]    = _c(row, "EMAIL SENT DATE","Email sent date","LAST EMAIL SENT DATE")
    r["Response"]           = _c(row, "Any Responses","replies","RESPONSES",
                                  "Email response","Email reply","Instagram Response","Linkedin reply")
    r["Notes"]              = _c(row, "comments","REmarks","notes","FOLLOW UP")
    r["Sales Person"]       = _c(row, "sales_person","SALES PERSON")
    r["Source"]             = source

    sf = source.lower()
    if category:
        r["Category"] = category
    elif "bathroom" in sf:
        r["Category"] = "Bathroom"
    elif "kitchen" in sf and "bath" in sf:
        r["Category"] = "KBB"
    elif "interior" in sf or "architect" in sf:
        r["Category"] = "Interior"
    elif "kitchen" in sf:
        r["Category"] = "Kitchen"
    elif "golden" in sf:
        r["Category"] = "KBB"

    return r

# ── Load all sources ───────────────────────────────────────────────────────────
COUNTRY_MAP = {
    "australia":"Australia","austria":"Austria","belgium":"Belgium",
    "france":"France","germany":"Germany","ireland":"Ireland",
    "netherlands":"Netherlands","spain":"Spain","sweden":"Sweden",
    "uk":"UK","usa":"USA","kbsa":"UK","cyncly":"UK","compusoft":"Multiple",
    "bathroom retailers uk":"UK","only bathrooms uk":"UK",
    "hubspot":"Multiple","golden customers":"Multiple","architects":"UK",
}

def load_csv(path):
    for enc in ("utf-8","latin-1","cp1252"):
        try:
            return pd.read_csv(path, dtype=str, encoding=enc, on_bad_lines="skip")
        except Exception:
            pass
    return pd.DataFrame()

def load_xlsx(path, sheets=None):
    try:
        xl  = pd.ExcelFile(path)
        out = []
        for s in (sheets or xl.sheet_names):
            try:
                df = xl.parse(s, dtype=str)
                if not df.empty:
                    out.append((s, df))
            except Exception:
                pass
        return out
    except Exception:
        return []

def load_all():
    records     = []   # all leads
    golden_keys = set()

    # ── Data Master.xlsx ──────────────────────────────────────────────────────
    cat_map = {"Kitchen":"Kitchen","KB":"KBB","Kitchen & Bathroom":"KBB","Bathroom":"Bathroom"}
    for sheet, df in load_xlsx(os.path.join(BASE,"Data Master.xlsx"),
                                ["Kitchen","KB","Kitchen & Bathroom","Bathroom"]):
        cat = cat_map.get(sheet,"Kitchen")
        for _, row in df.iterrows():
            lead = row_to_lead(row.to_dict(), f"Data Master/{sheet}", country="UK", category=cat)
            if lead["Company Name"]:
                records.append(lead)

    # ── Data Master (Old).xlsx ────────────────────────────────────────────────
    for sheet, df in load_xlsx(os.path.join(BASE,"Data Master ( Old ).xlsx"),["Sheet1"]):
        for _, row in df.iterrows():
            lead = row_to_lead(row.to_dict(), "Data Master (Old)", category="Kitchen")
            if lead["Company Name"]:
                records.append(lead)

    # ── Lead Master.xlsx ──────────────────────────────────────────────────────
    for sheet, df in load_xlsx(os.path.join(BASE,"Lead Master.xlsx")):
        for _, row in df.iterrows():
            lead = row_to_lead(row.to_dict(), f"Lead Master/{sheet}")
            if lead["Company Name"]:
                records.append(lead)

    # ── CAD leads CSVs ────────────────────────────────────────────────────────
    for fname in ("CAD leads(Customers).csv","CAD leads(Manish Leads).csv","CAD leads(Reps).csv"):
        path = os.path.join(BASE, fname)
        if not os.path.exists(path):
            continue
        df = load_csv(path)
        for _, row in df.iterrows():
            lead = row_to_lead(row.to_dict(), fname.replace(".csv",""))
            if lead["Company Name"]:
                records.append(lead)
                if "customer" in fname.lower():
                    golden_keys.add(_norm(lead["Company Name"]))

    # ── Old/ CSVs ─────────────────────────────────────────────────────────────
    for path in sorted(glob.glob(os.path.join(OLD_DIR,"*.csv"))):
        fname = os.path.basename(path)
        key   = re.sub(r"MASTER KITCHEN STUDIOS LIST\(|\)\.csv","",fname,flags=re.I).strip().lower()
        country = next((v for k,v in COUNTRY_MAP.items() if k in key), None)
        cat     = "Bathroom" if "bathroom" in fname.lower() else (
                  "Interior" if "architect" in fname.lower() else "Kitchen")
        df = load_csv(path)
        if df.empty:
            continue
        for _, row in df.iterrows():
            lead = row_to_lead(row.to_dict(), fname, country=country, category=cat)
            if lead["Company Name"]:
                records.append(lead)
                if "golden" in fname.lower():
                    golden_keys.add(_norm(lead["Company Name"]))

    return records, golden_keys


# ── Round-trip: read team edits from the existing Excel ───────────────────────
OUT_FILE = os.path.join(BASE, "CAD Illustrators - Master Lead Sheet.xlsx")

def load_team_edits() -> dict:
    """
    Reads the existing output Excel (if present) and returns a dict of
    normalised_company_name → {field: value} for every non-empty field.
    Reads BOTH the Master sheet (full field set) AND the Missing Data sheet
    (where the team fills in phone/email/website), so edits from either sheet
    survive the next rebuild.
    """
    if not os.path.exists(OUT_FILE):
        return {}
    edits: dict = {}
    try:
        xl = pd.ExcelFile(OUT_FILE)
    except Exception as e:
        print(f"  Warning: could not open existing Excel for round-trip: {e}")
        return {}

    contact_fields = {
        "Phone","Email","Website","LinkedIn (Owner)","LinkedIn (Company)",
        "Instagram","Facebook","Owner / Director","City","Postcode",
        "Full Address","Country","Brand(s)","Category","Outreach Status",
        "Email Sent Date","Follow-up Date","Response","Notes","Sales Person",
        "Tier","Score",
    }

    for sheet_title in xl.sheet_names:
        if not any(k in sheet_title for k in ("Master","Missing")):
            continue
        try:
            df = xl.parse(sheet_title, header=1, dtype=str)
        except Exception:
            continue
        df.columns = [str(c).strip() for c in df.columns]
        name_col = next((c for c in df.columns if "Company" in c), None)
        if not name_col:
            continue
        is_missing_sheet = "Missing" in sheet_title
        for _, row in df.iterrows():
            name = str(row.get(name_col) or "").strip()
            if not name or name.lower() in JUNK:
                continue
            key = _norm(name)
            if key not in edits:
                edits[key] = {}
            for col in contact_fields:
                if col in df.columns:
                    v = str(row.get(col) or "").strip()
                    if v and v.lower() not in JUNK:
                        # Missing Data sheet edits are newer → take priority
                        if is_missing_sheet or col not in edits[key]:
                            edits[key][col] = v

    print(f"  Team edits loaded: {len(edits):,} leads with saved field values")
    return edits


def apply_team_edits(records: list, edits: dict) -> list:
    """Merge team edits (from the previous Excel) into the current records."""
    if not edits:
        return records
    merged = 0
    for rec in records:
        key = _norm(rec.get("Company Name",""))
        if key in edits:
            for field, value in edits[key].items():
                if not rec.get(field):
                    rec[field] = value
                    merged += 1
    # Add leads that exist only in the old Excel (team-entered directly)
    existing_keys = {_norm(r.get("Company Name","")) for r in records}
    for key, fields in edits.items():
        if key not in existing_keys and fields.get("Company Name"):
            new_rec = {c: None for c in LEAD_COLS}
            new_rec.update(fields)
            new_rec["Source"] = new_rec.get("Source") or "team_edit"
            records.append(new_rec)
    if merged:
        print(f"  Applied {merged:,} team-edited field values from previous Excel")
    return records


def deduplicate(records):
    seen = {}   # normalised_name → index
    out  = []
    for rec in records:
        name = rec.get("Company Name") or ""
        key  = _norm(name)
        if not key:
            continue
        if key in seen:
            ex = out[seen[key]]
            for col in ("Phone","Email","Website","Owner / Director",
                        "LinkedIn (Owner)","LinkedIn (Company)",
                        "Instagram","Facebook","City","Postcode",
                        "Full Address","Country","Brand(s)"):
                if not ex.get(col) and rec.get(col):
                    ex[col] = rec[col]
        else:
            seen[key] = len(out)
            out.append(rec)
    return out

def load_job_leads():
    if not os.path.exists(JOB_CSV):
        return []
    df = load_csv(JOB_CSV)
    jobs = []
    for _, row in df.iterrows():
        row = {str(k): v for k, v in row.to_dict().items()}
        company = _c(row, "company","Company")
        if not company:
            continue
        jobs.append({
            "Found Date":    _c(row,"found_date"),
            "Tier":          _c(row,"tier"),
            "Score":         _c(row,"score"),
            "Job Title":     _c(row,"title"),
            "Company":       company,
            "Showroom Name": _c(row,"showroom_name"),
            "Location":      _c(row,"location"),
            "Salary":        _c(row,"salary"),
            "Is Recruiter":  _c(row,"is_recruiter"),
            "Opening Line":  _c(row,"opening_line"),
            "Job URL":       _c(row,"url"),
            "Source":        _c(row,"source"),
        })
    return jobs

# ── Sheet writers ──────────────────────────────────────────────────────────────

def write_header_row(ws, row_num, cols, bg=DARK, fg=GOLD):
    for ci, col in enumerate(cols, 1):
        header_cell(ws, row_num, ci, col, bg=bg, fg=fg)
    ws.row_dimensions[row_num].height = 26

def write_data_rows(ws, records, cols, start_row=2,
                    tier_col=None, highlight_missing=False):
    tier_colors = {"HOT":HOT_CLR,"WARM":WARM_CLR,"WATCH":WATCH_CLR}
    for ri, rec in enumerate(records, start=start_row):
        bg = LIGHT_BG if ri % 2 == 0 else ALT_BG
        for ci, col in enumerate(cols, 1):
            val  = rec.get(col)
            wrap = col in ("Full Address","Notes","Opening Line")
            c    = data_cell(ws, ri, ci, val, bg=bg, wrap=wrap)

        # Colour Tier cell
        if tier_col:
            parts = str(rec.get(tier_col) or "").split()
            tier_val = parts[0].upper() if parts else ""
            if tier_val in tier_colors:
                tc = ws.cell(row=ri, column=cols.index(tier_col)+1)
                tc.fill = fill(tier_colors[tier_val])
                tc.font = font(bold=True, color="FFFFFF")

        # Highlight rows missing key fields
        if highlight_missing:
            missing = not rec.get("Email") and not rec.get("Phone") and not rec.get("Website")
            if missing:
                for ci in range(1, len(cols)+1):
                    ws.cell(row=ri, column=ci).fill = fill(RED_LIGHT)

def build_leads_sheet(ws, records, title="All Leads", tab_color=DARK,
                      cols=None, widths=None, tier_col="Tier",
                      highlight_missing=False, hidden=False):
    cols   = cols   or LEAD_COLS
    widths = widths or COL_WIDTHS

    # Banner
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    banner = ws.cell(row=1, column=1, value=f"CAD Illustrators  —  {title}  ({len(records):,} leads)")
    banner.fill = fill(DARK)
    banner.font = font(bold=True, color=GOLD, size=12)
    banner.alignment = align("center")
    ws.row_dimensions[1].height = 30

    write_header_row(ws, 2, cols)
    set_col_widths(ws, cols, widths)
    ws.freeze_panes = "A3"

    write_data_rows(ws, records, cols, start_row=3,
                    tier_col=tier_col, highlight_missing=highlight_missing)

    # Outreach status dropdown
    if "Outreach Status" in cols:
        add_status_dropdown(ws, cols.index("Outreach Status")+1, 3, len(records)+10)

    # Auto-filter on header row
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"
    ws.sheet_properties.tabColor = tab_color

    if hidden:
        ws.sheet_state = "hidden"

# ── Summary sheet ──────────────────────────────────────────────────────────────
def write_summary(ws, all_records, golden_count, job_count, missing_count):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD

    def hdr(r, c, v, bg=DARK, fg=GOLD, sz=11, cols=1):
        if cols > 1:
            ws.merge_cells(start_row=r, start_column=c,
                           end_row=r, end_column=c+cols-1)
        cell = ws.cell(r, c, v)
        cell.fill = fill(bg); cell.font = font(True, fg, sz)
        cell.alignment = align("center")
        return cell

    def val(r, c, v, bg=LIGHT_BG, bold=False, fg="111111"):
        cell = ws.cell(r, c, v)
        cell.fill = fill(bg); cell.font = font(bold, fg)
        cell.alignment = align("center"); cell.border = border()
        return cell

    for col, w in [("A",30),("B",16),("C",30),("D",16),("E",2)]:
        ws.column_dimensions[col].width = w

    # Title
    hdr(1, 1, "CAD Illustrators — Lead Intelligence Master Sheet", DARK, GOLD, 14, cols=4)
    ws.row_dimensions[1].height = 38

    # KPI row
    kpis = [
        ("Total Leads", len(all_records), DARK),
        ("⭐ Golden", golden_count, GOLD),
        ("⚠️  Missing Data", missing_count, HOT_CLR),
        ("💼 Job Leads", job_count, WATCH_CLR),
    ]
    for ci, (label, count, bg) in enumerate(kpis, 1):
        ws.merge_cells(start_row=2, start_column=ci, end_row=2, end_column=ci)
        hdr(2, ci, f"{label}\n{count:,}", bg, GOLD if bg != GOLD else DARK, 11)
        ws.row_dimensions[2].height = 48
        ws.cell(2, ci).alignment = Alignment(horizontal="center", vertical="center",
                                              wrap_text=True)

    # By Category
    ws.row_dimensions[4].height = 22
    hdr(4, 1, "By Category", cols=2); hdr(4, 3, "By Country", cols=2)
    cats = Counter(r.get("Category") or "Unknown" for r in all_records)
    cnts = Counter(r.get("Country")  or "Unknown" for r in all_records)
    for i, ((cat, cc), (cty, cn)) in enumerate(
            zip(sorted(cats.items(), key=lambda x:-x[1]),
                sorted(cnts.items(),  key=lambda x:-x[1])), start=5):
        bg = LIGHT_BG if i%2 else ALT_BG
        val(i,1,cat,bg); val(i,2,cc,bg)
        val(i,3,cty,bg); val(i,4,cn,bg)

    # Data quality
    qrow = max(5+len(cats), 5+len(cnts)) + 2
    hdr(qrow, 1, "Data Quality", cols=2)
    total = len(all_records) or 1
    for label, count in [
        ("Has Email",    sum(1 for r in all_records if r.get("Email"))),
        ("Has Phone",    sum(1 for r in all_records if r.get("Phone"))),
        ("Has Website",  sum(1 for r in all_records if r.get("Website"))),
        ("Has Owner",    sum(1 for r in all_records if r.get("Owner / Director"))),
        ("Has LinkedIn", sum(1 for r in all_records if r.get("LinkedIn (Owner)"))),
    ]:
        qrow += 1
        bg = LIGHT_BG if qrow%2 else ALT_BG
        val(qrow, 1, label, bg)
        val(qrow, 2, f"{count/total*100:.0f}%  ({count:,})", bg)

# ── Missing data sheet ─────────────────────────────────────────────────────────
MISSING_COLS = [
    # Read-only context
    "Company Name", "Category", "Country", "City", "Postcode", "Full Address",
    # Team fills these ↓
    "Phone", "Email", "Website",
    "LinkedIn (Owner)", "LinkedIn (Company)", "Instagram", "Facebook",
    "Owner / Director",
    # Tracking
    "Updated By", "Notes",
]
MISSING_WIDTHS = {
    "Company Name":34, "Category":13, "Country":11, "City":15, "Postcode":10,
    "Full Address":36, "Phone":20, "Email":30, "Website":30,
    "LinkedIn (Owner)":30, "LinkedIn (Company)":30,
    "Instagram":26, "Facebook":26, "Owner / Director":24,
    "Updated By":16, "Notes":32,
}
# Columns the team needs to fill (highlighted green in header)
FILL_COLS = {"Phone","Email","Website","LinkedIn (Owner)","LinkedIn (Company)",
             "Instagram","Facebook","Owner / Director","Updated By","Notes"}
# Columns that are read-only context (greyed header)
READONLY_COLS = {"Company Name","Category","Country","City","Postcode","Full Address"}

def write_missing(ws, records):
    missing = [r for r in records
               if not r.get("Email") and not r.get("Phone") and not r.get("Website")]
    total   = len(records)
    done    = total - len(missing)
    pct     = int(done / total * 100) if total else 0

    num_cols = len(MISSING_COLS)
    span     = get_column_letter(num_cols)

    # Row 1: title banner
    ws.merge_cells(f"A1:{span}1")
    b = ws.cell(1, 1, f"Missing Data  —  {len(missing):,} leads still need contact info"
                      f"  |  Progress: {done:,} / {total:,} complete ({pct}%)")
    b.fill = fill(HOT_CLR); b.font = font(True,"FFFFFF",12)
    b.alignment = align("center"); ws.row_dimensions[1].height = 30

    # Row 2: instructions
    ws.merge_cells(f"A2:{span}2")
    inst = ws.cell(2, 1,
        "TEAM INSTRUCTIONS: Fill in Phone, Email or Website for each company. "
        "Green columns = fill these in.  Grey columns = read-only (do not edit).  "
        "Save the file, then re-run build_master.py — filled leads disappear from this list.")
    inst.fill = fill(MID); inst.font = font(italic=True, color="CCCCCC", size=10)
    inst.alignment = align("center"); ws.row_dimensions[2].height = 22

    # Row 3: progress bar (visual)
    ws.merge_cells(f"A3:{span}3")
    bar_filled = int(num_cols * pct / 100) or 1
    bar_text   = f"{'█' * bar_filled}{'░' * (num_cols - bar_filled)}  {pct}% complete"
    prog = ws.cell(3, 1, bar_text)
    prog.fill = fill(GREEN if pct == 100 else (WATCH_CLR if pct > 50 else HOT_CLR))
    prog.font = font(bold=True, color="FFFFFF", size=10)
    prog.alignment = align("center"); ws.row_dimensions[3].height = 18

    # Row 4: column headers — green for fill cols, grey for readonly
    for ci, col in enumerate(MISSING_COLS, 1):
        if col in FILL_COLS:
            bg, fg = GREEN, DARK
        else:
            bg, fg = "607070", "DDDDDD"
        header_cell(ws, 4, ci, col, bg=bg, fg=fg)
    ws.row_dimensions[4].height = 26

    set_col_widths(ws, MISSING_COLS, MISSING_WIDTHS)
    ws.freeze_panes = "A5"

    # Data rows
    for ri, rec in enumerate(missing, start=5):
        bg = AMBER if ri % 2 == 0 else RED_LIGHT
        for ci, col in enumerate(MISSING_COLS, 1):
            val  = rec.get(col) if col in LEAD_COLS else None
            wrap = col in ("Full Address","Notes")
            c    = data_cell(ws, ri, ci, val, bg=bg, wrap=wrap)
            # Dim read-only columns slightly
            if col in READONLY_COLS:
                c.font = font(color="777777", italic=True)

    ws.auto_filter.ref = f"A4:{span}4"
    ws.sheet_properties.tabColor = HOT_CLR
    return len(missing)

# ── Team Input sheet ───────────────────────────────────────────────────────────
TEAM_COLS   = ["Company Name","City / Area","Website (if known)","Notes / Context",
               "Added By","Status"]
TEAM_WIDTHS = {"Company Name":36,"City / Area":20,"Website (if known)":30,
               "Notes / Context":40,"Added By":18,"Status":20}

def write_team_input(ws):
    ws.merge_cells("A1:F1")
    t = ws.cell(1,1,"✏️   Team Input  —  Add company name here. Agent auto-fills everything else.")
    t.fill = fill(GREEN); t.font = font(True,DARK,12)
    t.alignment = align("center"); ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    s = ws.cell(2,1,"Only 'Company Name' is required. City/Area helps if the name is common. "
                    "DO NOT edit the Status column — the agent updates it automatically.")
    s.fill = fill(MID); s.font = font(italic=True,color="AAAAAA",size=10)
    s.alignment = align("center"); ws.row_dimensions[2].height = 20

    for ci,(col,w) in enumerate(TEAM_WIDTHS.items(), 1):
        c = ws.cell(3,ci,col)
        c.fill = fill(GOLD if col=="Company Name" else MID)
        c.font = font(True, DARK if col=="Company Name" else "CCCCCC")
        c.alignment = align("center"); c.border = border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 26
    ws.freeze_panes = "A4"

    for ri in range(4,254):
        bg = LIGHT_BG if ri%2==0 else ALT_BG
        for ci in range(1,7):
            c = ws.cell(ri,ci)
            c.fill = fill(bg); c.font = font(); c.alignment = align(); c.border = border()

    add_status_dropdown(ws, 6, 4, 253,
        '"Pending,Enriched,Drafted,Email Sent,Connected"')
    ws.auto_filter.ref = "A3:F3"
    ws.sheet_properties.tabColor = GREEN

# ── Entrypoint ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Loading all source files…")
    raw, golden_keys = load_all()
    print(f"  Raw records: {len(raw):,}")

    print("Loading team edits from previous Excel…")
    edits = load_team_edits()
    raw   = apply_team_edits(raw, edits)

    print("Deduplicating…")
    all_leads = deduplicate(raw)
    print(f"  Unique leads: {len(all_leads):,}")

    # Sort: UK first, then by country, then company name
    all_leads.sort(key=lambda r:(
        0 if (r.get("Country") or "").upper() in ("UK","UNITED KINGDOM") else 1,
        r.get("Country") or "ZZZ",
        (r.get("Company Name") or "").lower(),
    ))

    # Subsets
    showroom = [r for r in all_leads
                if (r.get("Country") or "").upper() in ("UK","UNITED KINGDOM","")
                and r.get("Category") in ("Kitchen","Bathroom","KBB","Fitter","Interior","Bedroom",None)]
    golden   = [r for r in all_leads
                if _norm(r.get("Company Name","")) in golden_keys]
    missing  = [r for r in all_leads
                if not r.get("Email") and not r.get("Phone") and not r.get("Website")]
    job_leads = load_job_leads()

    print(f"  Showroom leads : {len(showroom):,}")
    print(f"  Golden leads   : {len(golden):,}")
    print(f"  Missing data   : {len(missing):,}")
    print(f"  Job leads      : {len(job_leads):,}")

    print("Building workbook…")
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active; ws_sum.title = "📊 Summary"
    write_summary(ws_sum, all_leads, len(golden), len(job_leads), len(missing))

    # ── Sheet 2: Master (all data) ────────────────────────────────────────────
    ws_master = wb.create_sheet("📋 Master")
    build_leads_sheet(ws_master, all_leads, title="Master — All Leads", tab_color=DARK)

    # ── Sheet 3: Showroom Lists ───────────────────────────────────────────────
    ws_show = wb.create_sheet("🏠 Showroom Lists")
    build_leads_sheet(ws_show, showroom, title="Showroom Lists", tab_color=WATCH_CLR)

    # ── Sheet 4: Job Leads ────────────────────────────────────────────────────
    ws_jobs = wb.create_sheet("💼 Job Leads")
    ws_jobs.sheet_properties.tabColor = WARM_CLR
    ws_jobs.merge_cells(f"A1:{get_column_letter(len(JOB_COLS))}1")
    b = ws_jobs.cell(1,1,f"CAD Illustrators  —  Job Leads  ({len(job_leads):,} roles)")
    b.fill = fill(DARK); b.font = font(True,GOLD,12)
    b.alignment = align("center"); ws_jobs.row_dimensions[1].height = 30
    write_header_row(ws_jobs, 2, JOB_COLS, bg=WARM_CLR, fg="FFFFFF")
    set_col_widths(ws_jobs, JOB_COLS, JOB_WIDTHS)
    ws_jobs.freeze_panes = "A3"
    write_data_rows(ws_jobs, job_leads, JOB_COLS, start_row=3, tier_col="Tier")
    ws_jobs.auto_filter.ref = f"A2:{get_column_letter(len(JOB_COLS))}2"

    # ── Sheet 5: Golden Leads (hidden) ────────────────────────────────────────
    ws_gold = wb.create_sheet("⭐ Golden Leads")
    build_leads_sheet(ws_gold, golden, title="Golden Leads — High Value",
                      tab_color=GOLD, hidden=True)

    # ── Sheet 6: Missing Data ─────────────────────────────────────────────────
    ws_miss = wb.create_sheet("⚠️ Missing Data")
    write_missing(ws_miss, all_leads)

    # ── Sheet 7: Team Input ───────────────────────────────────────────────────
    ws_team = wb.create_sheet("✏️ Team Input")
    write_team_input(ws_team)

    wb.save(OUT_FILE)
    print(f"\nSaved: {OUT_FILE}")
    print(f"  Summary | Master ({len(all_leads):,}) | Showroom ({len(showroom):,})")
    print(f"  Jobs ({len(job_leads):,}) | Golden ({len(golden):,}) [hidden]")
    print(f"  Missing ({len(missing):,}) | Team Input (template)")
