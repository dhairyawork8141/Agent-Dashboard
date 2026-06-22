"""
excel_sync.py — Pulls company names from the team's live SharePoint Excel
('Manish Master sheet', ALL DATA sheet, Kitchen Retailers column) and syncs
new entries into Supabase as team_excel leads for the contact agent to enrich.

Deduplicates against:
  1. Existing Supabase leads (by normalised company name)
  2. Historical Data Analyst files (CSV/XLSX in Data Analist/)

Runs as the first step of contact-agent.yml so newly added companies are
enriched + drafted within the same run.
"""
import csv
import glob
import logging
import os
import re
import sys
from io import BytesIO

import openpyxl
import requests

log = logging.getLogger("excel_sync")

# ── SharePoint / OneDrive ──────────────────────────────────────────────────────
ONEDRIVE_USER = "abhay@cadillustrators.co.uk"
FILE_NAME     = "Manish Master sheet (1).xlsx"
SHEET_NAME    = "ALL DATA"
COL_HEADER    = "Kitchen Retailers"

# ── Env vars (same secrets the contact agent already uses) ─────────────────────
TENANT_ID     = os.environ.get("OAUTH_TENANT_ID", "")
CLIENT_ID     = os.environ.get("OAUTH_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("OAUTH_CLIENT_SECRET", "")
SUPABASE_URL  = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY  = os.environ.get("SUPABASE_SERVICE_KEY", "")
AGENT_ID      = os.environ.get("AGENT_ID", "")

# ── Path to historical data (relative to this file → repo root/Data Analist) ──
_HERE         = os.path.dirname(os.path.abspath(__file__))
DATA_ANALYST_DIR = os.path.join(_HERE, "..", "Data Analist")

# Column names (case-insensitive, stripped) used across Data Analyst files
_COMPANY_COLS = {
    "kitchen retailers", "bathroom retailers", "company name",
    "company", "name", "kitchens retailers name", "showroom/business name",
    "company name                      ",   # padded variant in one CSV
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _normalise(name: str) -> str:
    """Lowercase + strip all non-alphanumeric for fuzzy matching."""
    return re.sub(r"[^a-z0-9]", "", name.lower())


def _clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return None if s.lower() in ("", "nan", "none", "n/a", "na", "-") else s


# ── Microsoft Graph API ────────────────────────────────────────────────────────

def _graph_token() -> str:
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    r = requests.post(url, data={
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default",
        "grant_type":    "client_credentials",
    }, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def _download_excel(token: str) -> BytesIO:
    url = (f"https://graph.microsoft.com/v1.0/users/{ONEDRIVE_USER}"
           f"/drive/root:/{FILE_NAME}:/content")
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    r.raise_for_status()
    log.info("Downloaded %.1f KB from SharePoint", len(r.content) / 1024)
    return BytesIO(r.content)


# ── Parse the SharePoint Excel ─────────────────────────────────────────────────

def _parse_excel_names(excel_bytes: BytesIO) -> list[str]:
    wb = openpyxl.load_workbook(excel_bytes, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    col_idx = None
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value and COL_HEADER.lower() in str(cell.value).lower():
            col_idx = cell.column
            break
    if col_idx is None:
        raise ValueError(f"Column '{COL_HEADER}' not found in sheet '{SHEET_NAME}'")

    names = []
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        v = _clean(row[0].value)
        if v:
            names.append(v)
    wb.close()
    log.info("Parsed %d company names from SharePoint Excel", len(names))
    return names


# ── Load existing names for dedup ──────────────────────────────────────────────

def _supabase_names() -> set[str]:
    if not SUPABASE_URL or not SUPABASE_KEY:
        return set()
    try:
        r = requests.get(
            f"{SUPABASE_URL.rstrip('/')}/rest/v1/leads",
            headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
            params={"select": "company,showroom_name", "limit": "10000"},
            timeout=30,
        )
        r.raise_for_status()
        out = set()
        for row in r.json():
            for field in ("company", "showroom_name"):
                v = _clean(row.get(field))
                if v:
                    out.add(_normalise(v))
        log.info("Loaded %d existing names from Supabase", len(out))
        return out
    except Exception as e:
        log.warning("Failed to load Supabase names: %s", e)
        return set()


def _data_analyst_names() -> set[str]:
    """Extract company names from all CSV/XLSX files in the Data Analyst folder."""
    da_dir = os.path.abspath(DATA_ANALYST_DIR)
    if not os.path.isdir(da_dir):
        log.warning("Data Analyst dir not found at %s — skipping historical dedup", da_dir)
        return set()

    names: set[str] = set()

    # ── CSVs ──
    for path in glob.glob(os.path.join(da_dir, "**", "*.csv"), recursive=True):
        try:
            with open(path, encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue
                col = next(
                    (h for h in reader.fieldnames
                     if h and h.strip().lower() in _COMPANY_COLS),
                    None,
                )
                if not col:
                    continue
                for row in reader:
                    v = _clean(row.get(col))
                    if v:
                        names.add(_normalise(v))
        except Exception as e:
            log.debug("Skipping CSV %s: %s", os.path.basename(path), e)

    # ── Excel files ──
    for path in glob.glob(os.path.join(da_dir, "*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                if not row1:
                    continue
                headers = [str(h).strip().lower() if h else "" for h in row1[0]]
                col_idx = next(
                    (i for i, h in enumerate(headers) if h in _COMPANY_COLS), None
                )
                if col_idx is None:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    v = _clean(row[col_idx] if col_idx < len(row) else None)
                    if v:
                        names.add(_normalise(v))
            wb.close()
        except Exception as e:
            log.debug("Skipping Excel %s: %s", os.path.basename(path), e)

    log.info("Loaded %d historical names from Data Analyst files", len(names))
    return names


# ── Supabase upsert ────────────────────────────────────────────────────────────

def _upsert(companies: list[str]) -> int:
    if not companies or not SUPABASE_URL or not SUPABASE_KEY:
        return 0
    payload = [{
        "agent_id":     AGENT_ID or None,
        "external_key": f"excel_{_normalise(name)}",
        "company":      name,
        "showroom_name": name,
        "source":       "team_excel",
        "is_recruiter": False,
        # tier/enriched_at/draft_status left null — contact agent sets them
    } for name in companies]
    try:
        r = requests.post(
            f"{SUPABASE_URL.rstrip('/')}/rest/v1/leads",
            headers={
                "apikey":        SUPABASE_KEY,
                "Authorization": f"Bearer {SUPABASE_KEY}",
                "Content-Type":  "application/json",
                "Prefer":        "resolution=merge-duplicates,return=minimal",
            },
            params={"on_conflict": "external_key"},
            json=payload,
            timeout=30,
        )
        r.raise_for_status()
        return len(companies)
    except Exception as e:
        log.error("Upsert failed: %s", e)
        return 0


# ── Main ───────────────────────────────────────────────────────────────────────

def sync() -> int:
    """Returns number of new companies inserted into Supabase."""
    if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET]):
        log.warning("M365 OAuth creds missing — skipping Excel sync (set OAUTH_TENANT_ID/CLIENT_ID/CLIENT_SECRET)")
        return 0

    log.info("Fetching '%s' from SharePoint…", FILE_NAME)
    try:
        token = _graph_token()
        excel_bytes = _download_excel(token)
    except requests.HTTPError as e:
        log.error("SharePoint download failed: %s — check Files.Read.All permission", e)
        return 0

    raw_names = _parse_excel_names(excel_bytes)

    # Dedup within the sheet itself
    seen_keys: dict[str, str] = {}
    for name in raw_names:
        key = _normalise(name)
        if key and key not in seen_keys:
            seen_keys[key] = name
    unique_names = list(seen_keys.values())
    log.info("%d unique names after in-sheet dedup (removed %d dupes)",
             len(unique_names), len(raw_names) - len(unique_names))

    # Load known names from Supabase + Data Analyst files
    known = _supabase_names() | _data_analyst_names()

    new_companies = [n for n in unique_names if _normalise(n) not in known]
    skipped = len(unique_names) - len(new_companies)
    log.info("%d new | %d already known (Supabase + historical data)",
             len(new_companies), skipped)

    if not new_companies:
        log.info("Nothing new to sync.")
        return 0

    count = _upsert(new_companies)
    log.info("Synced %d new companies → Supabase (source=team_excel)", count)
    return count


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        stream=sys.stdout,
    )
    result = sync()
    print(f"\nExcel sync complete — {result} new companies queued for enrichment.")
